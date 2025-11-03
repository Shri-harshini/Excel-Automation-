from openpyxl import load_workbook

wb = load_workbook("Cola.xlsx")
ws = wb.active

# Create or get 'Summary' sheet
if 'Summary' in wb.sheetnames:
    summary = wb['Summary']
else:
    summary = wb.create_sheet("Summary")

total = 0
highest = 0
count = 0
top_name = ""

for idx, row in enumerate(ws.iter_rows(values_only=True)):
    if idx == 0:
        continue  # Skip header

    if not row or len(row) < 3:
        continue  # Skip empty or short rows

    name, region, sales = row[:3]  # Only take first 3 values

    if not isinstance(sales, (int, float)):
        continue  # Skip if sales is not a number

    total += sales
    count += 1

    if sales > highest:
        highest = sales
        top_name = name

summary["A1"] = "Total Sales"
summary["B1"] = total

summary["A2"] = "Average Sales"
summary["B2"] = total / count if count else 0

summary["A3"] = "Top Salesperson"
summary["B3"] = top_name

wb.save("Cola.xlsx")
print("✅ Sales summary saved to 'Cola.xlsx'")
